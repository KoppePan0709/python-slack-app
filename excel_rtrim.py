import openpyxl
SHEET_NAME = 'VPC'
wb=openpyxl.load_workbook('/Users/yuichi.masutani@ibm.com/Downloads/VPC.xlsx')
sheet = wb[SHEET_NAME]

sheet_max_row = sheet.max_row
sheet_max_column = sheet.max_column  # 数え始め 1

for row in range(1, sheet_max_row + 1):
    for column in range(1,sheet_max_column + 1):
        tmp = str(sheet.cell(row,column).value)
        sheet.cell(row,column).value = tmp.rstrip()


wb.save('/Users/yuichi.masutani@ibm.com/Downloads/VPC.xlsx')